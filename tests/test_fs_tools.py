import base64
import os
import time

import pytest

from filesystem_supertool2 import fs_tools


@pytest.fixture
def workdir(tmp_path):
    fs_tools.set_allowed_directories([str(tmp_path)])
    yield str(tmp_path)
    fs_tools.set_allowed_directories([])


def test_validate_path_rejects_relative(workdir):
    with pytest.raises(ValueError):
        fs_tools.validate_path("relative/path.txt")
    with pytest.raises(ValueError):
        fs_tools.validate_path("~/file.txt")


def test_validate_path_rejects_outside_allowed_directories(workdir):
    with pytest.raises(ValueError, match="Access denied"):
        fs_tools.validate_path("/etc/passwd", must_exist=False)


def test_validate_path_allows_subdirectory(workdir):
    sub = os.path.join(workdir, "a", "b")
    os.makedirs(sub)
    # should not raise
    fs_tools.validate_path(sub, must_exist=True)


def test_validate_path_no_allowed_directories_configured():
    fs_tools.set_allowed_directories([])
    with pytest.raises(ValueError, match="no allowed directories"):
        fs_tools.validate_path("/tmp/whatever", must_exist=False)


def test_validate_path_must_exist_false_allows_deep_missing_path(workdir):
    # Walks up to the nearest *existing* ancestor (workdir itself here), not just
    # the immediate parent -- so deep not-yet-created paths validate fine, matching
    # write_file/create_directory's mkdir_parents="create all missing levels" contract.
    deep = os.path.join(workdir, "a", "b", "c", "file.txt")
    resolved = fs_tools.validate_path(deep, must_exist=False)
    assert resolved == os.path.normpath(deep)


def test_validate_path_must_exist_false_rejects_symlinked_ancestor_escape(workdir, tmp_path_factory):
    # The requested path's literal string is inside the allowed root, but an ancestor
    # directory component is itself a symlink pointing outside it -- must still be
    # rejected once that ancestor is resolved, not just checked textually.
    outside = str(tmp_path_factory.mktemp("outside"))
    linked_dir = os.path.join(workdir, "linked")
    os.symlink(outside, linked_dir)
    not_yet_existing = os.path.join(linked_dir, "new_subdir", "file.txt")
    with pytest.raises(ValueError, match="Access denied"):
        fs_tools.validate_path(not_yet_existing, must_exist=False)


def test_validate_path_symlink_escape_rejected(workdir, tmp_path_factory):
    outside = str(tmp_path_factory.mktemp("outside"))
    target = os.path.join(outside, "secret.txt")
    with open(target, "w") as f:
        f.write("secret")
    link = os.path.join(workdir, "escape_link")
    os.symlink(target, link)
    with pytest.raises(ValueError, match="symlink target outside"):
        fs_tools.validate_path(link, must_exist=True)


def test_write_read_roundtrip(workdir):
    p = os.path.join(workdir, "a.txt")
    fs_tools.write_file(p, "hello\nworld\n")
    assert fs_tools.read_file(p) == "hello\nworld\n"


def test_write_file_overwrite_false(workdir):
    p = os.path.join(workdir, "a.txt")
    fs_tools.write_file(p, "one")
    with pytest.raises(FileExistsError):
        fs_tools.write_file(p, "two", overwrite=False)
    assert fs_tools.read_file(p) == "one"


def test_write_file_mkdir_parents(workdir):
    p = os.path.join(workdir, "nested", "dir", "a.txt")
    fs_tools.write_file(p, "x", mkdir_parents=True)
    assert fs_tools.read_file(p) == "x"


def test_read_file_binary_roundtrip(workdir):
    p = os.path.join(workdir, "bin.dat")
    raw = bytes(range(256))
    with open(p, "wb") as f:
        f.write(raw)
    encoded = fs_tools.read_file(p, binary=True)
    assert base64.b64decode(encoded) == raw


def test_read_file_head_tail_mutually_exclusive(workdir):
    p = os.path.join(workdir, "a.txt")
    fs_tools.write_file(p, "1\n2\n3\n")
    with pytest.raises(ValueError):
        fs_tools.read_file(p, head=1, tail=1)


def test_read_file_head_and_tail(workdir):
    p = os.path.join(workdir, "a.txt")
    fs_tools.write_file(p, "1\n2\n3\n4\n5\n")
    assert fs_tools.read_file(p, head=2) == "1\n2\n"
    assert fs_tools.read_file(p, tail=2) == "4\n5\n"


def test_read_file_line_numbers_with_tail(workdir):
    p = os.path.join(workdir, "a.txt")
    fs_tools.write_file(p, "1\n2\n3\n4\n5\n")
    out = fs_tools.read_file(p, tail=2, line_numbers=True)
    lines = out.splitlines()
    assert lines[0].strip().startswith("4|")
    assert lines[1].strip().startswith("5|")


def test_read_file_size_limit(workdir, monkeypatch):
    p = os.path.join(workdir, "big.txt")
    fs_tools.write_file(p, "x")
    monkeypatch.setattr(fs_tools, "MAX_READ_FILE_BYTES", 0)
    with pytest.raises(ValueError):
        fs_tools.read_file(p)
    # head/tail should still work despite the "file too big" condition
    assert fs_tools.read_file(p, head=1) == "x"


def test_edit_file_replaces_all_occurrences(workdir):
    p = os.path.join(workdir, "a.txt")
    fs_tools.write_file(p, "foo bar foo baz foo")
    result = fs_tools.edit_file(p, "foo", "qux")
    assert result["occurrences_replaced"] == 3
    assert fs_tools.read_file(p) == "qux bar qux baz qux"
    assert os.path.exists(p + ".bak")


def test_edit_file_dry_run_does_not_modify(workdir):
    p = os.path.join(workdir, "a.txt")
    fs_tools.write_file(p, "foo")
    result = fs_tools.edit_file(p, "foo", "bar", dry_run=True)
    assert result["dry_run"] is True
    assert fs_tools.read_file(p) == "foo"
    assert not os.path.exists(p + ".bak")


def test_edit_file_not_found_raises(workdir):
    p = os.path.join(workdir, "a.txt")
    fs_tools.write_file(p, "foo")
    with pytest.raises(ValueError):
        fs_tools.edit_file(p, "notfound", "x")


def test_edit_file_no_backup(workdir):
    p = os.path.join(workdir, "a.txt")
    fs_tools.write_file(p, "foo")
    fs_tools.edit_file(p, "foo", "bar", backup=False)
    assert not os.path.exists(p + ".bak")


def test_append_file_no_auto_newline(workdir):
    p = os.path.join(workdir, "a.txt")
    fs_tools.write_file(p, "foo")
    fs_tools.append_file(p, "bar")
    assert fs_tools.read_file(p) == "foobar"


def test_append_file_creates_missing(workdir):
    p = os.path.join(workdir, "new.txt")
    fs_tools.append_file(p, "hello")
    assert fs_tools.read_file(p) == "hello"


def test_delete_file_requires_confirm(workdir):
    p = os.path.join(workdir, "a.txt")
    fs_tools.write_file(p, "x")
    with pytest.raises(ValueError):
        fs_tools.delete_file(p, confirm=False)
    assert os.path.exists(p)
    fs_tools.delete_file(p, confirm=True)
    assert not os.path.exists(p)


def test_copy_file(workdir):
    src = os.path.join(workdir, "src.txt")
    dst = os.path.join(workdir, "dst.txt")
    fs_tools.write_file(src, "content")
    fs_tools.copy_file(src, dst)
    assert fs_tools.read_file(dst) == "content"
    assert fs_tools.read_file(src) == "content"


def test_copy_file_destination_must_not_be_directory(workdir):
    src = os.path.join(workdir, "src.txt")
    fs_tools.write_file(src, "content")
    os.makedirs(os.path.join(workdir, "destdir"))
    with pytest.raises(ValueError):
        fs_tools.copy_file(src, os.path.join(workdir, "destdir"))


def test_copy_file_overwrite_false(workdir):
    src = os.path.join(workdir, "src.txt")
    dst = os.path.join(workdir, "dst.txt")
    fs_tools.write_file(src, "content")
    fs_tools.write_file(dst, "existing")
    with pytest.raises(FileExistsError):
        fs_tools.copy_file(src, dst, overwrite=False)


def test_move_file_rename(workdir):
    src = os.path.join(workdir, "old.txt")
    dst = os.path.join(workdir, "new.txt")
    fs_tools.write_file(src, "content")
    fs_tools.move_file(src, dst)
    assert not os.path.exists(src)
    assert fs_tools.read_file(dst) == "content"


def test_create_directory_idempotent(workdir):
    d = os.path.join(workdir, "sub", "dir")
    fs_tools.create_directory(d)
    fs_tools.create_directory(d)  # should not raise
    assert os.path.isdir(d)


def test_delete_directory_requires_empty_without_force(workdir):
    d = os.path.join(workdir, "sub")
    os.makedirs(d)
    fs_tools.write_file(os.path.join(d, "f.txt"), "x")
    with pytest.raises(OSError):
        fs_tools.delete_directory(d, confirm=True)
    fs_tools.delete_directory(d, confirm=True, force=True)
    assert not os.path.exists(d)


def test_list_directory_hidden_default_true(workdir):
    fs_tools.write_file(os.path.join(workdir, "visible.txt"), "x")
    fs_tools.write_file(os.path.join(workdir, ".hidden.txt"), "x")
    result = fs_tools.list_directory(workdir)
    names = {e["name"] for e in result["entries"]}
    assert "visible.txt" in names
    assert ".hidden.txt" in names


def test_list_directory_hidden_false(workdir):
    fs_tools.write_file(os.path.join(workdir, "visible.txt"), "x")
    fs_tools.write_file(os.path.join(workdir, ".hidden.txt"), "x")
    result = fs_tools.list_directory(workdir, hidden=False)
    names = {e["name"] for e in result["entries"]}
    assert "visible.txt" in names
    assert ".hidden.txt" not in names


def test_list_directory_count_only(workdir):
    fs_tools.write_file(os.path.join(workdir, "a.txt"), "x")
    os.makedirs(os.path.join(workdir, "sub"))
    result = fs_tools.list_directory(workdir, count_only=True)
    assert result["files"] == 1
    assert result["directories"] == 1
    assert result["total"] == 2


def test_list_directory_detail_levels(workdir):
    fs_tools.write_file(os.path.join(workdir, "a.txt"), "hello")
    minimal = fs_tools.list_directory(workdir, detail="minimal")["entries"][0]
    assert set(minimal.keys()) == {"name"}
    standard = fs_tools.list_directory(workdir, detail="standard")["entries"][0]
    assert {"name", "type", "size"} <= set(standard.keys())
    full = fs_tools.list_directory(workdir, detail="full")["entries"][0]
    assert {"modified", "permissions"} <= set(full.keys())


def _make_tree(base):
    os.makedirs(os.path.join(base, "src", "sub"))
    os.makedirs(os.path.join(base, "node_modules", "pkg"))
    fs_tools.write_file(os.path.join(base, "root.py"), "x")
    fs_tools.write_file(os.path.join(base, "src", "a.py"), "x")
    fs_tools.write_file(os.path.join(base, "src", "sub", "b.py"), "x")
    fs_tools.write_file(os.path.join(base, "src", "a.txt"), "x")
    fs_tools.write_file(os.path.join(base, "node_modules", "pkg", "index.js"), "x")


def test_find_files_recursive_glob(workdir):
    _make_tree(workdir)
    result = fs_tools.find_files(workdir, "**/*.py")
    rels = sorted(os.path.relpath(m, workdir) for m in result["matches"])
    assert rels == sorted(["root.py", "src/a.py", "src/sub/b.py"])


def test_find_files_excludes_node_modules_by_default(workdir):
    _make_tree(workdir)
    result = fs_tools.find_files(workdir, "**/*.js")
    assert result["matches"] == []


def test_find_files_max_depth_reports_boundary_dirs(workdir):
    _make_tree(workdir)
    # max_depth=1: base directory only -- top-level entries visible, no recursion
    result = fs_tools.find_files(workdir, "*", max_depth=1, type="directory", exclude="")
    names = {os.path.basename(m) for m in result["matches"]}
    assert names == {"src", "node_modules"}


def test_find_files_max_depth_2_sees_one_level_down(workdir):
    _make_tree(workdir)
    result = fs_tools.find_files(workdir, "**/*.py", max_depth=2)
    rels = sorted(os.path.relpath(m, workdir) for m in result["matches"])
    # src/sub/b.py is two levels down from base -- should NOT appear with max_depth=2
    assert "src/sub/b.py" not in rels
    assert "src/a.py" in rels
    assert "root.py" in rels


def test_find_files_type_filter(workdir):
    _make_tree(workdir)
    result = fs_tools.find_files(workdir, "*", type="directory", exclude="")
    assert all(os.path.isdir(m) for m in result["matches"])


def test_find_files_count_only(workdir):
    _make_tree(workdir)
    result = fs_tools.find_files(workdir, "**/*.py")
    assert "matches" not in fs_tools.find_files(workdir, "**/*.py", count_only=True)
    assert fs_tools.find_files(workdir, "**/*.py", count_only=True)["count"] == len(result["matches"])


def test_find_files_brace_expansion(workdir):
    fs_tools.write_file(os.path.join(workdir, "a.js"), "x")
    fs_tools.write_file(os.path.join(workdir, "a.ts"), "x")
    fs_tools.write_file(os.path.join(workdir, "a.py"), "x")
    result = fs_tools.find_files(workdir, "*.{js,ts}")
    names = sorted(os.path.basename(m) for m in result["matches"])
    assert names == ["a.js", "a.ts"]


def test_grep_basic_literal_case_insensitive(workdir):
    p = os.path.join(workdir, "a.txt")
    fs_tools.write_file(p, "Hello World\nhello again\nGoodbye\n")
    result = fs_tools.grep(workdir, "hello")
    assert result["total"] == 2


def test_grep_case_sensitive(workdir):
    p = os.path.join(workdir, "a.txt")
    fs_tools.write_file(p, "Hello World\nhello again\n")
    result = fs_tools.grep(workdir, "hello", case_sensitive=True)
    assert result["total"] == 1


def test_grep_regex(workdir):
    p = os.path.join(workdir, "a.txt")
    fs_tools.write_file(p, "def foo():\ndef bar():\nclass Baz:\n")
    result = fs_tools.grep(workdir, r"^def \w+", regex=True)
    assert result["total"] == 2


def test_grep_skips_binary_files(workdir):
    p = os.path.join(workdir, "bin.dat")
    with open(p, "wb") as f:
        f.write(b"hello\x00world")
    result = fs_tools.grep(workdir, "hello")
    assert result["total"] == 0


def test_grep_count_only(workdir):
    p = os.path.join(workdir, "a.txt")
    fs_tools.write_file(p, "match\nmatch\nno\n")
    result = fs_tools.grep(workdir, "match", count_only=True)
    assert result["counts"][p] == 2
    assert "matches" not in result


def test_grep_single_file(workdir):
    p = os.path.join(workdir, "a.txt")
    fs_tools.write_file(p, "match\n")
    result = fs_tools.grep(p, "match")
    assert result["total"] == 1


def test_stat_file_and_directory(workdir):
    p = os.path.join(workdir, "a.txt")
    fs_tools.write_file(p, "hello")
    file_stat = fs_tools.stat(p)
    assert file_stat["type"] == "file"
    assert file_stat["size"] == 5
    dir_stat = fs_tools.stat(workdir)
    assert dir_stat["type"] == "directory"


def test_exists(workdir):
    p = os.path.join(workdir, "a.txt")
    assert fs_tools.exists(p)["exists"] is False
    fs_tools.write_file(p, "x")
    assert fs_tools.exists(p)["exists"] is True
    assert fs_tools.exists(p, type="directory")["exists"] is False
    assert fs_tools.exists(workdir, type="directory")["exists"] is True


def test_checksum(workdir):
    p = os.path.join(workdir, "a.txt")
    fs_tools.write_file(p, "hello")
    import hashlib

    expected = hashlib.sha256(b"hello").hexdigest()
    assert fs_tools.checksum(p)["digest"] == expected
    assert fs_tools.checksum(p, algorithm="md5")["digest"] == hashlib.md5(b"hello").hexdigest()


def test_compare_files_identical(workdir):
    a = os.path.join(workdir, "a.txt")
    b = os.path.join(workdir, "b.txt")
    fs_tools.write_file(a, "same\n")
    fs_tools.write_file(b, "same\n")
    result = fs_tools.compare_files(a, b)
    assert result["identical"] is True
    assert result["diff"] == ""


def test_compare_files_different(workdir):
    a = os.path.join(workdir, "a.txt")
    b = os.path.join(workdir, "b.txt")
    fs_tools.write_file(a, "one\n")
    fs_tools.write_file(b, "two\n")
    result = fs_tools.compare_files(a, b)
    assert result["identical"] is False
    assert "-one" in result["diff"]
    assert "+two" in result["diff"]


def test_patch_file_apply_and_reverse(workdir):
    a = os.path.join(workdir, "a.txt")
    b = os.path.join(workdir, "b.txt")
    fs_tools.write_file(a, "one\n")
    fs_tools.write_file(b, "two\n")
    diff = fs_tools.compare_files(a, b)["diff"]

    fs_tools.patch_file(a, diff)
    assert fs_tools.read_file(a) == "two\n"
    assert os.path.exists(a + ".bak")

    fs_tools.patch_file(a, diff, reverse=True, backup=False)
    assert fs_tools.read_file(a) == "one\n"


def test_touch_creates_and_updates(workdir):
    p = os.path.join(workdir, "a.txt")
    result = fs_tools.touch(p)
    assert result["created"] is True
    assert os.path.exists(p)
    mtime_before = os.path.getmtime(p)
    time.sleep(0.01)
    result = fs_tools.touch(p)
    assert result["created"] is False
    assert os.path.getmtime(p) >= mtime_before


def test_tail_follow_incremental(workdir):
    p = os.path.join(workdir, "log.txt")
    fs_tools.write_file(p, "line1\n")
    first = fs_tools.tail_follow(p)
    assert first["content"] == ""  # first call establishes cursor, per spec

    fs_tools.append_file(p, "line2\n")
    second = fs_tools.tail_follow(p)
    assert second["content"] == "line2\n"

    third = fs_tools.tail_follow(p)
    assert third["content"] == ""


def test_tail_follow_reset(workdir):
    p = os.path.join(workdir, "log.txt")
    fs_tools.write_file(p, "line1\n")
    fs_tools.tail_follow(p)
    fs_tools.append_file(p, "line2\n")
    fs_tools.tail_follow(p, reset=True)
    assert fs_tools.tail_follow(p)["content"] == ""


def test_tree_structure_and_exclude(workdir):
    _make_tree(workdir)
    result = fs_tools.tree(workdir)
    assert "node_modules" not in result["tree"]
    assert "src" in result["tree"]
    assert "root.py" in result["tree"]


def test_tree_max_depth(workdir):
    _make_tree(workdir)
    result = fs_tools.tree(workdir, max_depth=1, exclude="")
    assert "root.py" in result["tree"]
    assert "sub" not in result["tree"]


def test_read_lines_range(workdir):
    p = os.path.join(workdir, "a.txt")
    fs_tools.write_file(p, "1\n2\n3\n4\n5\n")
    assert fs_tools.read_lines(p, 2, 4, line_numbers=False) == "2\n3\n4\n"


def test_read_lines_to_end(workdir):
    p = os.path.join(workdir, "a.txt")
    fs_tools.write_file(p, "1\n2\n3\n")
    assert fs_tools.read_lines(p, 2, -1, line_numbers=False) == "2\n3\n"


def test_read_lines_with_numbers(workdir):
    p = os.path.join(workdir, "a.txt")
    fs_tools.write_file(p, "a\nb\n")
    out = fs_tools.read_lines(p, 1, 2)
    assert out.splitlines()[0].strip().startswith("1|")


def test_get_hostname_strips_domain():
    result = fs_tools.get_hostname()
    assert "." not in result["hostname"]
    assert result["fqdn"]


def test_read_document_unsupported_extension(workdir):
    p = os.path.join(workdir, "a.csv")
    fs_tools.write_file(p, "a,b\n1,2\n")
    with pytest.raises(ValueError):
        fs_tools.read_document(p)


def test_read_document_docx(workdir):
    import docx

    p = os.path.join(workdir, "a.docx")
    d = docx.Document()
    d.add_paragraph("Hello from docx")
    d.save(p)
    text = fs_tools.read_document(p)
    assert "Hello from docx" in text


def test_read_document_xlsx(workdir):
    import openpyxl

    p = os.path.join(workdir, "a.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "hello"
    ws["B1"] = "world"
    wb.save(p)
    text = fs_tools.read_document(p)
    assert "Sheet1" in text
    assert "hello" in text
    assert "world" in text
